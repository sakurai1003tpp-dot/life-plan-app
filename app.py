import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_insurance_simulation():
    wb = openpyxl.Workbook()

    # 1. ワークシートの設定
    ws_dash = wb.active
    ws_dash.title = "ダッシュボード"
    ws_detail = wb.create_sheet(title="詳細シミュレーション")

    # グリッド線の表示
    ws_dash.views.sheetView[0].showGridLines = True
    ws_detail.views.sheetView[0].showGridLines = True

    # スタイルの定義
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid") # ダークスレートグレー
    header_font = Font(name="Meiryo", size=11, bold=True, color="FFFFFF")
    
    cell_font = Font(name="Meiryo", size=10)
    bold_cell_font = Font(name="Meiryo", size=10, bold=True)
    title_font = Font(name="Meiryo", size=16, bold=True, color="2F4F4F")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # --- ダッシュボードタブ ---
    ws_dash['B2'] = "ライフプラン・世帯収支ダッシュボード"
    ws_dash['B2'].font = title_font

    ws_dash['B4'] = "基本前提サマリー（二人分保険料・夫85歳死亡保険金反映）"
    ws_dash['B4'].font = bold_cell_font

    summary_params = [
        ("項目", "設定値", "備考"),
        ("夫の年齢（現在）", 40, "シミュレーション開始時点"),
        ("妻の年齢（現在）", 38, "シミュレーション開始時点"),
        ("夫の想定死亡年齢", 85, "この時点で保険金受給（掛け金二人分の半分を反映）"),
        ("世帯月額保険料", 15000, "二人分の保険料"),
        ("夫の死亡時保険金", 6000000, "二人分保険料ベースの保障の半分"),
    ]

    for row_idx, row_data in enumerate(summary_params, start=5):
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws_dash.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if row_idx == 5:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            else:
                if col_idx == 3 and isinstance(val, (int, float)):
                    if val > 1000:
                        cell.number_format = '¥#,##0'
                    else:
                        cell.number_format = '#,##0'

    # --- 詳細シミュレーションタブ (サイド詳細項目をデフォルトで折りたたみ) ---
    ws_detail['B2'] = "年次キャッシュフロー詳細（サイド詳細項目はデフォルトで折りたたみ）"
    ws_detail['B2'].font = title_font

    headers = [
        "年", "夫年齢", "妻年齢", 
        "世帯手取り収入（基本）", "死亡保険金受給", "世帯支出", "年間収支", 
        "[詳細] 夫給与", "[詳細] 妻給与", "[詳細] 夫年金", "[詳細] 妻年金", "[詳細] 保険料控除等"
    ]

    for col_idx, h in enumerate(headers, start=2):
        cell = ws_detail.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 46年分のシミュレーション（夫40歳〜85歳）
    start_h_age = 40
    start_w_age = 38

    for i in range(46):
        r = 5 + i
        h_age = start_h_age + i
        w_age = start_w_age + i
        is_death_year = (h_age == 85)
        
        death_benefit = 6000000 if is_death_year else 0
        
        ws_detail.cell(row=r, column=2, value=2026 + i).alignment = Alignment(horizontal="center")
        ws_detail.cell(row=r, column=3, value=h_age).alignment = Alignment(horizontal="center")
        ws_detail.cell(row=r, column=4, value=w_age).alignment = Alignment(horizontal="center")
        
        # 基本手取り収入
        net_inc = 5000000 if h_age < 65 else 2500000
        ws_detail.cell(row=r, column=5, value=net_inc)
        ws_detail.cell(row=r, column=5).number_format = '¥#,##0'
        
        ws_detail.cell(row=r, column=6, value=death_benefit)
        ws_detail.cell(row=r, column=6).number_format = '¥#,##0'
        
        expense = 4000000 if h_age < 85 else 3000000
        ws_detail.cell(row=r, column=7, value=expense)
        ws_detail.cell(row=r, column=7).number_format = '¥#,##0'
        
        # 年間収支計算式
        ws_detail.cell(row=r, column=8, value=f"=E{r}+F{r}-G{r}")
        ws_detail.cell(row=r, column=8).number_format = '¥#,##0'
        ws_detail.cell(row=r, column=8).font = bold_cell_font
        
        # 詳細列（I列〜M列）: グループ化して折りたたむ対象
        ws_detail.cell(row=r, column=9, value=3500000 if h_age < 65 else 0) # 夫給与
        ws_detail.cell(row=r, column=10, value=1500000 if w_age < 60 else 0) # 妻給与
        ws_detail.cell(row=r, column=11, value=0 if h_age < 65 else 1500000) # 夫年金
        ws_detail.cell(row=r, column=12, value=0 if w_age < 65 else 1000000) # 妻年金
        ws_detail.cell(row=r, column=13, value=-180000) # 保険料（二人分年額）
        
        for c in range(9, 14):
            ws_detail.cell(row=r, column=c).number_format = '¥#,##0'
            
        for c in range(2, 14):
            ws_detail.cell(row=r, column=c).border = thin_border
            ws_detail.cell(row=r, column=c).font = cell_font

    # 列IからMをグループ化し、デフォルトで非表示（折りたたみ）にする
    ws_detail.column_dimensions.group('I', 'M', hidden=True)

    # 列幅の自動調整
    for ws in [ws_dash, ws_detail]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    file_name = "insurance_simulation_report.xlsx"
    wb.save(file_name)
    return file_name

if __name__ == "__main__":
    generate_insurance_simulation()
